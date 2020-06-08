
#This program extracts data from a table in SQL within a date range specified by startdate and enddate variable.

import pandas as pd
import pyodbc as podbc
import datetime

from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl as xl

# cursor = sql_conn.cursor()
# cursor.execute('SELECT CIF_NO, CLIENT_TYPE FROM MMS.CLIENT')

startdate = datetime.datetime(2019, 4, 1)
enddate = datetime.datetime(2019, 4, 30)
report_type = 'test_report'
firstsheet = True

wb = xl.Workbook()
writer = pd.ExcelWriter('test_report.xlsx')
# LMS 08
assert startdate <= enddate
while startdate <= enddate:

    wstitle = str(startdate.strftime('%m-%d-%y'))

    sql_conn = podbc.connect('DRIVER={SQL Server};'
                             'SERVER=test_server;'
                             'DATABASE=test;'
                             'UID=username;'
                             'PWD=pword;'
                             'Trusted_Connection=no')

# region LMS_08
    if report_type == 'test_report':
        report_name = 'test_report.xlsx'

        query = "SELECT  " \
                "B.PRODUCT_NAME AS 'LOAN TYPE',  " \
                "B.PRODUCT_SHORT_NAME + '-' + (CASE WHEN A.ACCOUNT_STATUS_CODE = DBO.GET_RPT_VALUE('LOAN_STATUS4')  " \
                "THEN 'NPL' WHEN A.ACCOUNT_STATUS_CODE = DBO.GET_RPT_VALUE('LOAN_STATUS1') THEN 'CURRENT' ELSE  " \
                "DBO.GET_DESC(A.ACCOUNT_STATUS_CODE,'LOAN_STATUS') END) + '-' + A.TRAN_CODE AS 'SHEET NAME', " \
                "DBO.GET_DESC(A.ACCOUNT_STATUS_CODE,'LOAN_STATUS') AS 'LOAN STATUS',  " \
                "A.TRAN_CODE AS 'TRAN CODE',  " \
                "A.BUSINESS_DATE AS 'TRAN DATE',  " \
                "A.TRAN_CODE_DESCRIPTION AS 'TRAN CODE DESCRIPTION'," \
                "A.BRANCH_CODE + '-' + D.BRANCH_NAME AS 'BRANCH OFFICE',  " \
                "A.BUSINESS_DATE AS 'TRANSACTION DATE',  " \
                "''''+ A.OR_NO AS OR_NO, " \
                "A.OR_DATE,  " \
                "A.ACCOUNT_NAME AS 'NAME', " + \
                "''''+ A.ACCOUNT_NO AS 'LOAN ACCOUNT ID',  " \
                "'''' +A.SERIAL_NO AS 'SERIAL NO', " + \
                "A.PAYJUR_CODE AS 'PAYJUR',  " \
                "A.LOAN_LEDGER_TOTAL_AMOUNT AS 'TOTAL PRINCIPAL AMOUNT',  " \
                "A.PAY_PERIOD AS 'PAY PERIOD',  " \
                "ISNULL(A.PRINCIPAL, 0) AS PRINCIPAL,  " \
                "CASE   " \
                "WHEN A.TRAN_CODE = '2624' THEN 0 " \
                "ELSE ISNULL(A.INTEREST, 0) " \
                "END AS INTEREST,  " \
                "CASE   " \
                "WHEN A.TRAN_CODE = '2624' THEN 0 " \
                "ELSE ISNULL(A.AIR, 0)" \
                "END AS AIR,  " \
                "ISNULL(A.PENALTY, 0) AS PENALTY, " \
                "ISNULL(A.COLLECTION_AGENT_FEE, 0) AS 'COLLECTION AGENT FEE',  " \
                "ISNULL(A.CCA_OVERPAYMENT, 0) AS 'CCA OVERPAYMENT',  " \
                "ISNULL(A.SD_OVERPAYMENT, 0) AS 'SD OVERPAYMENT', " + \
                "ISNULL(A.AP, 0) AS AP,  " \
                "ISNULL(A.AR, 0) AS AR,  " \
                "CASE " \
                "WHEN A.TRAN_CODE = '2711' AND E.PRODUCT_CODE = '001' THEN A.INTEREST " \
                "ELSE ISNULL(A.CCA, 0) " \
                "END AS CCA,  " \
                "CASE " \
                "WHEN A.TRAN_CODE = '2711' AND E.PRODUCT_CODE <> '001' THEN A.INTEREST "  \
                "ELSE ISNULL(A.SD, 0) " \
                "END AS SD,  " \
                "(  " \
                "ISNULL(A.PRINCIPAL, 0) +  " \
                "ISNULL(A.INTEREST, 0) +  " \
                "ISNULL(A.AIR, 0) +  " \
                "ISNULL(A.PENALTY, 0) +  " \
                "ISNULL(A.COLLECTION_AGENT_FEE, 0) + " \
                "ISNULL(A.AP, 0) +  " \
                "ISNULL(A.AR, 0) +  " \
                "ISNULL(A.CCA, 0) +  " \
                "ISNULL(A.SD, 0)) AS TOTAL,  " \
                "A.USER_NAME AS 'USER ID'  " \
                "FROM xxx.xxx_HISTORY_CONSOLIDATED A WITH(NOLOCK)  " \
                "LEFT JOIN xxx1.xxx.PRODUCT_DETAILS B WITH(NOLOCK) ON A.PRODUCT_CODE = B.PRODUCT_CODE " + \
                "LEFT JOIN xxx2.xxx.ACCOUNT_INFO C WITH(NOLOCK) ON C.ACCOUNT_NO = A.ACCOUNT_NO   " \
                "LEFT JOIN xxx3.BRANCH_DETAILS D WITH(NOLOCK) ON A.BRANCH_CODE = D.BRANCH_CODE  " \
                "LEFT JOIN xxx2.xxx.ACCOUNT_BALANCE E  WITH(NOLOCK) ON E.ACCOUNT_NO = A.DEPOSIT_ACCOUNT_NO " \
                "WHERE   " \
                "AND DATEDIFF(DAY,A.BUSINESS_DATE, '" + str(startdate) + "') = 0   " \
                "AND A.TRAN_CODE <> 'xxx'   "

# endregion

# region LMS_09
    elif report_type == 'test_report1':
        report_name = 'test_report1.xlsx'

        query = "SELECT " \
                "B.PRODUCT_NAME AS 'LOAN TYPE', " \
                "B.PRODUCT_SHORT_NAME + '-' + (CASE WHEN A.ACCOUNT_STATUS_CODE = DBO.GET_RPT_VALUE('LOAN_STATUS4') THEN 'NPL' WHEN A.ACCOUNT_STATUS_CODE = DBO.GET_RPT_VALUE('LOAN_STATUS1') THEN 'CURRENT' ELSE DBO.GET_DESC(A.ACCOUNT_STATUS_CODE,'LOAN_STATUS') END) + '-' + A.TRAN_CODE AS 'SHEET NAME', " \
                "DBO.GET_DESC(A.ACCOUNT_STATUS_CODE,'LOAN_STATUS') AS 'LOAN STATUS', " \
                "A.TRAN_CODE AS 'TRAN CODE', " \
                "A.BUSINESS_DATE AS 'TRAN DATE', " \
                "A.TRAN_CODE_DESCRIPTION AS 'TRAN CODE DESCRIPTION', " \
                "A.BRANCH_CODE + '-' + D.BRANCH_NAME AS 'BRANCH OFFICE', " \
                "A.BUSINESS_DATE AS 'TRANSACTION DATE', " \
                "''''+A.OR_NO AS OR_NO, " \
                "A.OR_DATE, " \
                "A.ACCOUNT_NAME AS 'NAME', " \
                "''''+A.ACCOUNT_NO AS 'LOAN ACCOUNT ID', " \
                "''''+A.SERIAL_NO AS 'SERIAL NO', " \
                "A.PAYJUR_CODE AS 'PAYJUR', " \
                "A.LOAN_LEDGER_TOTAL_AMOUNT AS 'TOTAL PRINCIPAL AMOUNT', " \
                "A.PAY_PERIOD AS 'PAY PERIOD', " \
                "ISNULL(A.PRINCIPAL, 0) AS PRINCIPAL, " \
                "ISNULL(A.INTEREST, 0) AS INTEREST, " \
                "ISNULL(A.AIR, 0) AS AIR, " \
                "ISNULL(A.PENALTY, 0) AS PENALTY, " \
                "ISNULL(A.COLLECTION_AGENT_FEE, 0) AS 'COLLECTION AGENT FEE', " \
                "ISNULL(A.CCA_OVERPAYMENT, 0) AS 'CCA OVERPAYMENT', " \
                "ISNULL(A.SD_OVERPAYMENT, 0) AS 'SD OVERPAYMENT', " \
                "ISNULL(A.AP, 0) AS AP, " \
                "ISNULL(A.AR, 0) AS AR, " \
                "ISNULL(A.CCA, 0) AS CCA, " \
                "ISNULL(A.SD, 0) AS SD, " \
                "(ISNULL(A.PRINCIPAL, 0) + " \
                "ISNULL(A.INTEREST, 0) + " \
                "ISNULL(A.AIR, 0) + " \
                "ISNULL(A.PENALTY, 0) + " \
                "ISNULL(A.COLLECTION_AGENT_FEE, 0) + " \
                "ISNULL(A.AP, 0) + " \
                "ISNULL(A.AR, 0) + " \
                "ISNULL(A.CCA, 0) + " \
                "ISNULL(A.SD, 0)) " \
                "AS TOTAL, " \
                "A.USER_NAME AS 'USER ID' " \
                "FROM xxx.xxx_HISTORY_CONSOLIDATED A WITH(NOLOCK) " \
                "LEFT JOIN xxx1.xxx.PRODUCT_DETAILS B WITH(NOLOCK) ON A.PRODUCT_CODE = B.PRODUCT_CODE " \
                "LEFT JOIN xxx2.xxx.ACCOUNT_INFO C WITH(NOLOCK) ON C.ACCOUNT_NO = A.ACCOUNT_NO " \
                "LEFT JOIN xxx3.BRANCH_DETAILS D WITH(NOLOCK) ON A.BRANCH_CODE = D.BRANCH_CODE " \
                "WHERE " \
                "A.USER_NAME IN (SELECT RPT_VALUE FROM RPT.REPORT_VARIABLE WHERE RPT_CODE LIKE 'SYSTEM_USERNAME[0123456789]%') " \
                "AND DATEDIFF(DAY,A.BUSINESS_DATE,'" + str(startdate) + "') = 0 " \
                "AND A.TRAN_CODE <> 'test' "


# endregion

# region LMS_10
    else:
        report_name = 'test_report2.xlsx'

        query = "SELECT " \
                "PD.PRODUCT_SHORT_NAME + '-' + IIF(LHC.ACCOUNT_STATUS_CODE = DBO.GET_RPT_VALUE('LOAN_STATUS4'),'NPL',DBO.GET_DESC(LHC.ACCOUNT_STATUS_CODE,'LOAN_STATUS')) AS 'SHEET NAME' " \
                ",DBO.GET_DESC(LHC.PRODUCT_CODE, 'LOAN_TYPE') AS 'LOAN TYPE' " \
                ",DBO.GET_DESC(LHC.ACCOUNT_STATUS_CODE, 'LOAN_STATUS') AS 'LOAN STATUS' " \
                ",LHC.TRAN_CODE AS 'TRAN CODE' " \
                ",LHC.TRAN_CODE_DESCRIPTION AS 'DESCRIPTION'	 " \
                ",COUNT(LHC.TRAN_CODE) AS 'NO OF RECORDS' " \
                ",SUM(IIF(ITEMS.TRAN_TYPE = DBO.get_val('TRAN_TYPE1') OR LHC.TRAN_CODE IN ('some values'), ISNULL(PRINCIPAL,0),0)) AS 'DEBIT' " \
                ",SUM(IIF(ITEMS.TRAN_TYPE = DBO.get_val('TRAN_TYPE2') OR LHC.TRAN_CODE IN ('some values') ,ISNULL(PRINCIPAL,0),0)) AS 'CREDIT' " \
                "FROM xxx.xxx1_HISTORY_CONSOLIDATED LHC WITH(NOLOCK) " \
                "	INNER JOIN xxx1.TRAN_ITEM_DETAILS ITEMS WITH(NOLOCK) " \
                "		ON LHC.TRAN_CODE = ITEMS.TRAN_CODE " \
                "	INNER JOIN xxx2.xxx.PRODUCT_DETAILS PD WITH(NOLOCK) " \
                "		ON LHC.PRODUCT_CODE = PD.PRODUCT_CODE " \
                "WHERE DATEDIFF(DAY,LHC.BUSINESS_DATE, '" + str(startdate) + "') = 0 " \
                "AND (ITEMS.TRAN_TYPE <> (SELECT RPT_VALUE FROM RPT.REPORT_VARIABLE WITH(NOLOCK) WHERE RPT_CODE = 'test_code') OR LHC.TRAN_CODE = '2658') " \
                "AND LHC.TRAN_CODE NOT IN ('test1') " \
                "AND PRINCIPAL > 0  " \
                "GROUP BY PD.PRODUCT_SHORT_NAME,LHC.PRODUCT_CODE, LHC.ACCOUNT_STATUS_CODE, LHC.TRAN_CODE, LHC.TRAN_CODE_DESCRIPTION "
# endregion

    print('Querying: %s ' % wstitle)
    df = pd.read_sql(sql=query, con=sql_conn)
    sql_conn.close()

    if df.empty:
        print('Skipped %s, no data.' % wstitle)
        startdate += datetime.timedelta(days=1)
        continue

    for col in df.select_dtypes(include=['datetime']):
        df[col] = df[col].dt.strftime('%m/%d/%Y')

    #if firstsheet is True:
    #   firstsheet = False
    #    ws = wb.active
    #   ws.title = wstitle
    #else:
    #    ws = wb.create_sheet(wstitle)  # insert at the end (default)
    # or
    # ws2 = wb.create_sheet("wstitle", 0)  # insert at first position

    print('Appending data to worksheet: %s ' % wstitle)
    #for r in dataframe_to_rows(df, index=False, header=True):
    #    ws.append(r)
    df.to_excel(writer, sheet_name=wstitle, index=False, header=True)

    print('Done with: %s ' % wstitle)
    startdate += datetime.timedelta(days=1)

#wb.save(report_name)
writer.save()
print('Report created:  %s.' % report_name)
